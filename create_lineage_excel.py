#!/usr/bin/env python3
"""
Create Comprehensive Data Lineage Excel

This script generates a detailed Excel workbook containing all data lineage
information for all 13 studies, with multiple sheets for easy analysis.
"""

import json
import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

LINEAGE_DIR = ".claude/YDATA_LOCAL_PACKAGE/lineage"
OUTPUT_FILE = "DATA_LINEAGE_COMPLETE.xlsx"

def load_all_lineage():
    """Load all lineage JSON files."""
    lineage_files = glob.glob(os.path.join(LINEAGE_DIR, "*_harmonization_report_v2.json"))

    all_lineage = {}
    for filepath in sorted(lineage_files):
        study_id = os.path.basename(filepath).replace('_harmonization_report_v2.json', '')
        with open(filepath, 'r') as f:
            all_lineage[study_id] = json.load(f)

    return all_lineage

def create_overview_sheet(all_lineage):
    """Create overview summary of all studies."""
    overview_data = []

    for study_id, lineage in sorted(all_lineage.items()):
        row = {
            'Study ID': study_id,
            'Harmonization Date': lineage.get('harmonization_date', 'N/A'),
            'Version': lineage.get('version', lineage.get('harmonization_info', {}).get('harmonization_version', 'N/A')),
            'Total Records': lineage.get('output_summary', lineage.get('dataset_summary', {})).get('total_records', 'N/A'),
            'Unique Patients': lineage.get('output_summary', lineage.get('dataset_summary', {})).get('unique_patients', 'N/A'),
            'Records per Patient': lineage.get('output_summary', lineage.get('dataset_summary', {})).get('records_per_patient', lineage.get('output_summary', lineage.get('dataset_summary', {})).get('records_per_patient_mean', 'N/A')),
            'Source Files Count': len(lineage.get('source_files_used', {})),
            'Variables with Lineage': len(lineage.get('data_lineage', {})),
        }
        overview_data.append(row)

    return pd.DataFrame(overview_data)

def create_variable_lineage_sheet(all_lineage):
    """Create comprehensive variable lineage table."""
    lineage_data = []

    for study_id, lineage in sorted(all_lineage.items()):
        data_lineage = lineage.get('data_lineage', {})

        for variable, info in sorted(data_lineage.items()):
            if isinstance(info, dict):
                # Extract sources
                sources = []
                if 'source' in info:
                    sources.append(info['source'])
                if 'primary_source' in info:
                    sources.append(f"{info['primary_source']} (primary)")
                if 'secondary_source' in info:
                    sources.append(f"{info['secondary_source']} (secondary)")

                source_text = "; ".join(sources) if sources else "Multiple sources"

                # Handle mappings
                mapping_text = ""
                if 'mapping' in info and isinstance(info['mapping'], dict):
                    mapping_items = [f"{k}→{v}" for k, v in list(info['mapping'].items())[:5]]
                    mapping_text = "; ".join(mapping_items)
                    if len(info['mapping']) > 5:
                        mapping_text += f" (and {len(info['mapping'])-5} more)"

                row = {
                    'Study ID': study_id,
                    'Variable Name': variable,
                    'Source(s)': source_text,
                    'Description': info.get('description', ''),
                    'Value Mappings': mapping_text,
                }
            else:
                row = {
                    'Study ID': study_id,
                    'Variable Name': variable,
                    'Source(s)': str(info),
                    'Description': '',
                    'Value Mappings': '',
                }

            lineage_data.append(row)

    return pd.DataFrame(lineage_data)

def create_source_files_sheet(all_lineage):
    """Create source files summary."""
    source_data = []

    for study_id, lineage in sorted(all_lineage.items()):
        source_files = lineage.get('source_files_used', {})

        for file_id, info in sorted(source_files.items()):
            row = {
                'Study ID': study_id,
                'File ID': file_id,
                'Description': info.get('description', 'N/A'),
                'Records': info.get('records', 'N/A'),
            }
            source_data.append(row)

    return pd.DataFrame(source_data)

def create_biomarker_coverage_sheet(all_lineage):
    """Create biomarker coverage summary."""
    biomarker_data = []

    for study_id, lineage in sorted(all_lineage.items()):
        coverage = lineage.get('biomarker_coverage', {})

        for biomarker, info in sorted(coverage.items()):
            if isinstance(info, dict):
                row = {
                    'Study ID': study_id,
                    'Biomarker': biomarker,
                    'Total Measurements': info.get('total_measurements', info),
                    'Unique Patients': info.get('unique_patients', 'N/A'),
                    'Measurements per Patient': info.get('measurements_per_patient', 'N/A'),
                }
            else:
                row = {
                    'Study ID': study_id,
                    'Biomarker': biomarker,
                    'Total Measurements': info,
                    'Unique Patients': 'N/A',
                    'Measurements per Patient': 'N/A',
                }
            biomarker_data.append(row)

    return pd.DataFrame(biomarker_data) if biomarker_data else None

def style_worksheet(ws, df, title):
    """Apply styling to worksheet."""
    # Set title
    ws.title = title[:31]  # Excel sheet name limit

    # Header styling
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Apply alternating row colors
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

def main():
    """Main execution."""
    print("="*70)
    print("CREATING COMPREHENSIVE DATA LINEAGE EXCEL")
    print("="*70)

    # Load all lineage data
    print("\nLoading lineage data...")
    all_lineage = load_all_lineage()
    print(f"✓ Loaded {len(all_lineage)} studies")

    # Create Excel workbook
    print("\nCreating Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Overview
    print("  • Creating Overview sheet...")
    overview_df = create_overview_sheet(all_lineage)
    ws_overview = wb.create_sheet("Overview")
    for r in dataframe_to_rows(overview_df, index=False, header=True):
        ws_overview.append(r)
    style_worksheet(ws_overview, overview_df, "Overview")

    # Sheet 2: Variable Lineage
    print("  • Creating Variable Lineage sheet...")
    lineage_df = create_variable_lineage_sheet(all_lineage)
    ws_lineage = wb.create_sheet("Variable Lineage")
    for r in dataframe_to_rows(lineage_df, index=False, header=True):
        ws_lineage.append(r)
    style_worksheet(ws_lineage, lineage_df, "Variable Lineage")

    # Sheet 3: Source Files
    print("  • Creating Source Files sheet...")
    source_df = create_source_files_sheet(all_lineage)
    ws_source = wb.create_sheet("Source Files")
    for r in dataframe_to_rows(source_df, index=False, header=True):
        ws_source.append(r)
    style_worksheet(ws_source, source_df, "Source Files")

    # Sheet 4: Biomarker Coverage (if available)
    biomarker_df = create_biomarker_coverage_sheet(all_lineage)
    if biomarker_df is not None and not biomarker_df.empty:
        print("  • Creating Biomarker Coverage sheet...")
        ws_biomarker = wb.create_sheet("Biomarker Coverage")
        for r in dataframe_to_rows(biomarker_df, index=False, header=True):
            ws_biomarker.append(r)
        style_worksheet(ws_biomarker, biomarker_df, "Biomarker Coverage")

    # Save workbook
    print(f"\nSaving {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)

    print(f"✓ Created {OUTPUT_FILE}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"  Total variables: {len(lineage_df)}")
    print(f"  Total source files: {len(source_df)}")

    print("\n" + "="*70)
    print("COMPLETE")
    print("="*70)

if __name__ == "__main__":
    main()
